import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from datetime import datetime
import time
import os

# -----------------------------
# 1️⃣  Get Amazon Price
# -----------------------------
def get_amazon_price(url):
    options = Options()
    options.add_argument("--headless=new")  # Run without opening a browser window
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    
    driver = webdriver.Chrome(options=options)
    driver.get(url)
    time.sleep(2)  # Wait for page to load
    
    price = None
    
    try:

       price = driver.find_element(By.ID, "priceblock_ourprice").text

    except:
        try:
            price = driver.find_element(By.ID, "priceblock_dealprice").text
        except:
            try:
                price = driver.find_element(By.CLASS_NAME, "a-price-whole").text
                cents = driver.find_element(By.CLASS_NAME, "a-price-fraction").text
                price = f"${price}.{cents}"
            except:
                print(f"⚠️  Price not found for URL: {url}")


    driver.quit()
    return price

# -----------------------------
# 2️⃣  Save Price to Excel
# -----------------------------
def save_price_to_excel(sku, price):
    date = datetime.today().strftime('%Y-%m-%d')
    output_file = "daily_prices.xlsx"
    new_data = pd.DataFrame([[date, sku, price]], columns=['Date', 'SKU', 'Price'])

    if os.path.exists(output_file):
        if os.path.getsize(output_file) == 0:
            os.remove(output_file)
            new_data.to_excel(output_file, index=False)
            print(f"✅ Created fresh Excel file and saved price for {sku}")
            return
            book = load_workbook(output_file)
        
        with pd.ExcelWriter(output_file, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            new_data.to_excel(writer, sheet_name='Sheet1', index=False, header=False,
        startrow=writer.sheets['Sheet1'].max_row)

        new_data.to_excel(writer, sheet_name='Sheet1', index=False, header=False,
                          startrow=writer.sheets['Sheet1'].max_row)
        writer.save()
    else:
        new_data.to_excel(output_file, index=False)
    print(f"✅ Saved {sku} price: {price}")

# -----------------------------
# 3️⃣  Compare Price Parity
# -----------------------------
def check_price_parity():
    if not os.path.exists("daily_prices.xlsx"):
        print("ℹ️ No previous price data found. Skipping parity check.")
    return
    df = pd.read_excel("daily_prices.xlsx")

    latest = df.groupby('SKU').tail(2).sort_values(by=['SKU','Date'])

    print("\n🔍 Price Parity Check:")
    for sku in latest['SKU'].unique():
        last_two = latest[latest['SKU'] == sku]['Price'].tolist()
        if len(last_two) == 2:
            yesterday, today = last_two
            status = "Constant"
            try:
                yesterday_val = float(yesterday.replace('$', '').replace(',', ''))
                today_val = float(today.replace('$', '').replace(',', ''))
                if today_val > yesterday_val:
                    status = "Increased"
                elif today_val < yesterday_val:
                    status = "Decreased"
            except:
                status = "Price format error"
            print(f"{sku}: {status}")

# -----------------------------
# 4️⃣  Main Function
# -----------------------------
def main():
    products = pd.read_excel("products.xlsx")
    for index, row in products.iterrows():
        sku = row['SKU']
        url = row['URL']
        price = get_amazon_price(url)
        if price:
            save_price_to_excel(sku, price)
    
    check_price_parity()

# -----------------------------
# 5️⃣  Run Script
# -----------------------------
if __name__ == "__main__":
    main()
